import json
import math

from PIL import Image

from PyQt5 import QtCore, QtWidgets, QtGui, Qt
import io
import folium
import openpyxl as opxl

from PyQt5.QtGui import QPixmap, QPainter, QPen, QBrush, QColor
from PyQt5.QtWebEngineWidgets import QWebEngineView
from os import listdir, path
from os.path import isfile, join
from interface.python_ui.map_ui import Ui_MapWindow
from PyQt5.QtWidgets import QFileDialog, QGraphicsRectItem, QTableWidget, QTableWidgetItem
from error import ErrorWindow



class PhotoViewer(QtWidgets.QGraphicsView):
    photoClicked = QtCore.pyqtSignal(QtCore.QPoint)

    def __init__(self, parent):
        super(PhotoViewer, self).__init__(parent)
        self._zoom = 0
        self._empty = True
        self._scene = QtWidgets.QGraphicsScene(self)

        self._photo = QtWidgets.QGraphicsPixmapItem()
        self._scene.addItem(self._photo)


        self.name_pic = ""
        self.scale_2 = 0.2
        self.bugRectItems = []
        self.setScene(self._scene)
        self.setTransformationAnchor(QtWidgets.QGraphicsView.AnchorUnderMouse)
        self.setResizeAnchor(QtWidgets.QGraphicsView.AnchorUnderMouse)

        #self.setFrameShape(QtWidgets.QFrame.NoFrame)

    def setName(self, name):
        #print("set")
        self.name_pic = name
        #self.update()

    def hasPhoto(self):
        return not self._empty

    def fitInView(self, scale=True):
        print("f-----")
        rect = QtCore.QRectF(self._photo.pixmap().rect())
        if not rect.isNull():
            self.setSceneRect(rect)
            if self.hasPhoto():
                unity = self.transform().mapRect(QtCore.QRectF(0, 0, 1, 1))
                print(unity)
                self.scale(1 / unity.width(), 1 / unity.height())
                viewrect = self.viewport().rect()
                print(viewrect)
                scenerect = self.transform().mapRect(rect)
                print(scenerect)
                factor = min(viewrect.width() / scenerect.width(),
                             viewrect.height() / scenerect.height())
                print(factor)
                self.scale(factor, factor)
            self._zoom = 0
        print("------")

    def setPhoto(self, pixmap=None):
        for i in range(len(self.bugRectItems)):
            self._scene.removeItem(self.bugRectItems[i])
        self._zoom = 0
        print("_1")
        if pixmap and not pixmap.isNull():
            print("_2")
            self._empty = False
            self.setDragMode(QtWidgets.QGraphicsView.ScrollHandDrag)
            print("_4")
            self._photo.setPixmap(pixmap)
            print("_6")
        else:
            print("_3")
            self._empty = True
            self.setDragMode(QtWidgets.QGraphicsView.NoDrag)
            print("_5")
            self._photo.setPixmap(QtGui.QPixmap())
            print("_7")
        self.fitInView()
        print("_-_")

    def wheelEvent(self, event):
        if self.hasPhoto():
            if event.angleDelta().y() > 0:
                factor = 1.25
                self._zoom += 1
            else:
                factor = 0.8
                self._zoom -= 1

            if self._zoom > 0:
                self.scale(factor, factor)
            elif self._zoom == 0:
                print("fitInView")
                self.fitInView()
            else:
                self._zoom = 0
        print(self._zoom)



    def set_rect_bugs(self):
        for i in range(len(self.bugRectItems)):
            self._scene.removeItem(self.bugRectItems[i])
        print("remove")
        print("draw")

        with open(self.name_pic[:-4] + ".json", "r") as read_file:
            data_now_pic = json.load(read_file)
        print(data_now_pic)
        print(len(data_now_pic["targets"]))
        for i in range(len(data_now_pic["targets"])):
            x = data_now_pic["targets"][i]["polygon"]
            print(data_now_pic["targets"][i]["class_id"])
            print(x)
            cords = []
            num = ''
            for j in x:
                if j.isdigit():
                    num += j
                else:
                    if num != '':
                        cords.append(int(num))
                        num = ''
            print(cords)

            w_s = cords[0]
            h_s = cords[1]
            w_e = cords[4] - cords[0]
            h_e = cords[5] - cords[1]
            print("paint zoom - ", self.scale_2)
            rect = self._scene.addRect(int(w_s), int(h_s), int(w_e), int(h_e))
            self.bugRectItems.append(rect)
            print("ok")
            if data_now_pic["targets"][i]["class_id"] == "1001":
                print("ok2")
                pen = QPen(QColor(233, 159, 94))
                pen.setWidth(2)
                rect.setPen(pen)
            elif data_now_pic["targets"][i]["class_id"] == "1002":
                print("ok3")
                pen = QPen(QColor(254,30,0))
                pen.setWidth(2)
                rect.setPen(pen)
            elif data_now_pic["targets"][i]["class_id"] == "1003":
                print("ok4")
                pen = QPen(QColor(244,16,194))
                pen.setWidth(2)
                rect.setPen(pen)
            elif data_now_pic["targets"][i]["class_id"] == "1004":
                print("ok5")
                pen = QPen(QColor(250,252,1))
                pen.setWidth(2)
                rect.setPen(pen)
            else:
                print("error")

            # print("paint")




class MapWindow(QtWidgets.QMainWindow, Ui_MapWindow):

    def __init__(self):
        super(MapWindow, self).__init__()
        print("init")
        self.setupUi(self)

        self.error = ErrorWindow()
        #self.label = Label()

        self.Language = ["name" , "name_ru"]
        self.lang = 0
        self.scale = 0.2
        self.mydir_map = ""
        self.name_now_pic = ""
        self.qpixmax = QPixmap("../output data/pictures/DJI_0415.jpg")
        self.masCoords = []
        self.few = 0
        self.many = 0
        self.all = "Все виды"
        self.statistics = {"1001": 0, "1002": 0, "1003": 0, "1004": 0, "square": 0}
        self.telemetry_path = ""
        self.viewer = PhotoViewer(self)
        self.verticalLayout_pic.addWidget(self.viewer)
        self.bugs_dd = {"1001": 0, "1002": 0, "1003": 0, "1004": 0}
        self.bugs_map = {"1001": 0, "1002": 0, "1003": 0, "1004": 0}

        self.add_functions()

    def setLang(self, lang):
        self.lang = lang - 1
    def setFew(self, few):
        self.few = few
        print(self.few)
    def setMany(self, many):
        self.many = many
        print(self.many)

    def add_functions(self):
        print("add")
        #self.map_coords(48.03089, 37.68103, "aboba")
        self.button_select_folder.clicked.connect(self.on_click_select_folder_map)
        self.comboBox_all_bugs.currentIndexChanged.connect(self.search_bugs_set)
        self.pushButton_delet_all.clicked.connect(self.dell_all)
        self.pushButton_telem.clicked.connect(self.setTelemetryPath)

    def setTelemetryPath(self):
        dialog = QFileDialog()
        self.telemetry_path = QFileDialog.getOpenFileName()[0]



    def dell_all(self):
        print("99")
        print(self.verticalLayout.count())
        for i in reversed(range(1, self.verticalLayout.count())):
            self.verticalLayout.itemAt(i).widget().setParent(None)
        self.comboBox_all_bugs.clear()
        for i in range(self.verticalLayout_map.count()):
            self.verticalLayout_map.itemAt(i).widget().setParent(None)
        self.qpixmax = QPixmap("D:/BUGS_Programm/bin/interface/imges/JILDU.jpg")
        self.viewer.setPhoto(self.qpixmax)
        self.label_name.setText("")
        self.textBrowser_Flight_statistic.setText("")
        self.textBrowser_disease_and_recom.setText("")
        print("ne sdelal tek kek xz chto tut doljilno bIdt'")

    def on_click_select_folder_map(self):
        print()
        dialog = QFileDialog()
        self.mydir_map = dialog.getExistingDirectory()
        self.do_buttons()

        print("mydir_map " + self.mydir_map)
        pass

    def do_buttons(self):
        print("ee")
        k = 0
        mypath = self.mydir_map
        onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
        wb = opxl.load_workbook(self.telemetry_path)
        sheet = wb['Лист1']



        for i in range(0, len(onlyfiles) - 1, 2):

            while True:
                if onlyfiles[i][:-4] == sheet['A' + str(k + 1)].value[:-4]:
                    number_in_telemetry = k
                    break
                k += 1
            x = onlyfiles[i]
            path = self.mydir_map + "/" + onlyfiles[i][:-4] + ".json"
            print(path)
            print(x)

            with open("info.json", "r") as read_file:
                data_of_info = json.load(read_file)
            with Image.open(self.mydir_map + "/" + onlyfiles[i][:-4] + ".jpg") as img:
                img.load()

            self.statistics["square"] += 4 * int(sheet['F' + str(k + 1)].value) ** 2 * math.tan(math.radians(data_of_info["Camera_angle"] / 2)) * math.tan(math.atan(img.height / img.width * math.tan(math.radians(data_of_info["Camera_angle"] / 2))))
            with open(path, "r") as read_file:
                data_now_pic = json.load(read_file)
            self.x = QtWidgets.QPushButton(self.scrollAreaWidgetContents)
            self.x.setLayoutDirection(QtCore.Qt.LeftToRight)
            self.x.setObjectName("x")
            self.x.setText(x)
            for j in range(len(data_now_pic["targets"])):
                if data_now_pic["targets"][j]["class_id"] == "1001":
                    self.statistics["1001"] += 1
                if data_now_pic["targets"][j]["class_id"] == "1002":
                    self.statistics["1002"] += 1
                if data_now_pic["targets"][j]["class_id"] == "1003":
                    self.statistics["1003"] += 1
                if data_now_pic["targets"][j]["class_id"] == "1004":
                    self.statistics["1004"] += 1

            if len(data_now_pic["targets"]) <= self.few:
                print("1")
                self.x.setStyleSheet('QPushButton {background-color: rgb(64,255,81)}')
            elif len(data_now_pic["targets"]) <= self.many and len(data_now_pic["targets"]) > self.few:
                print("2")
                self.x.setStyleSheet('QPushButton {background-color: rgb(255,246,91)}')
            elif len(data_now_pic["targets"]) > self.many:
                print("3")
                self.x.setStyleSheet('QPushButton {background-color: rgb(255,127,76)}')
            print("lll")
            self.verticalLayout.addWidget(self.x)
            print("eee")
            self.x.clicked.connect(lambda state, x = x: self.chek_img(x))
            print("bbb")
        print("всё нормас")
        print(self.statistics)
        self.textBrowser_Flight_statistic.setText("В этом полёте обнаружено:\n Bedbug evil turtle - " +  str(self.statistics["1001"]) + "\n Bread striped flea - " + str(self.statistics["1002"]) + "\n Bread beetle - " + str(self.statistics["1003"]) + "\n Meadow moth - " + str(self.statistics["1004"]) + "\n Расчетная площадь порытия фотографиями: " + str(self.statistics["square"]) + " м^2")

    def openErrorWindow(self, text):
        self.error.show()
        self.error.label_error.setText(text)
        self.error.buttonCloseErrorWindow.clicked.connect(self.go_back_map)

    def go_back_map(self):
        self.error.close()

    def chek_img(self, name):
        self.bugs_dd["1001"] = 0
        self.bugs_dd["1002"] = 0
        self.bugs_dd["1003"] = 0
        self.bugs_dd["1004"] = 0
        k = 0
        self.name_now_pic = name
        print(self.mydir_map + "/" + name)
        print(path.exists(self.mydir_map + "/" + name))
        print(self.mydir_map + "/" + name[:-4] + ".json")
        print(path.exists(self.mydir_map + "/" + name[:-4] + ".json"))
        if path.exists(self.mydir_map + "/" + name) and path.exists(self.mydir_map + "/" + name[:-4] + ".json"):
            wb = opxl.load_workbook(self.telemetry_path)
            sheet = wb['Лист1']
            while True:
                if name[:-4] == sheet['A' + str(k + 1)].value[:-4]:
                    number_in_telemetry = k
                    break
                k += 1
            x = sheet['B' +  str(number_in_telemetry + 1)].value
            y = sheet['C' +  str(number_in_telemetry + 1)].value
            name_pic = self.mydir_map + "/" + name

            print(x, y, name_pic, name)
            self.diseases_text(self.mydir_map + "/" + name[:-4] + ".json")
            self.qpixmax = QPixmap(name_pic)
            self.search_type_bugs_set(self.mydir_map + "/" + name[:-4] + ".json")
            self.viewer.setName(name_pic)
            self.map_coords(x, y, name)
            self.viewer.setPhoto(self.qpixmax)
            self.viewer.set_rect_bugs()
            print("^")
            self.label_name.setText(name)
        else:
            self.openErrorWindow("Set the path to the folder with images")


    def diseases_text(self, pathjson):
        text_diseases = ""
        wb = opxl.load_workbook("D:/BUGS_Programm/database/diseases base/base_diseases.xlsx")
        sheet = wb['Лист1']
        num_des = 0
        l = [wb['Лист1']
             ['a' + str(row)].value for row in range(2, wb['Лист1'].max_row + 1)
             if (wb['Лист1']['B' + str(row)].value) != None]
        with open(pathjson, "r") as read_file:
            data_now_pic = json.load(read_file)

        for i in range(len(data_now_pic["targets"])):
            if data_now_pic["targets"][i]["class_id"] == "1001":
                self.bugs_dd["1001"] += 1
            elif data_now_pic["targets"][i]["class_id"] == "1002":
                self.bugs_dd["1002"] += 1
            elif data_now_pic["targets"][i]["class_id"] == "1003":
                self.bugs_dd["1003"] += 1
            elif data_now_pic["targets"][i]["class_id"] == "1004":
                self.bugs_dd["1004"] += 1

        for i in range(len(l)):
            if int(sheet['C' + str(i + 2)].value) <= self.bugs_dd["100" + str(i + 1)]:
                num_des += 1
                u = i + 2
                if text_diseases != "":
                    text_diseases += "\n"
                text_diseases += str(num_des) + ") На данном снимке выявлена болезь под названием '" + sheet['D' + str(u)].value + "'\n Цвет полигона: " + sheet['K' + str(u)].value +\
                                 "\n  Лечение. \nОпрыскивать препаратом '" + sheet['E' + str(u)].value +\
                        "' \nНорма препарата л/га: " + sheet['H' + str(u)].value + " \nНорма рабочей жидкости л/га: " + sheet['I' + str(u)].value +\
                        " \nСпособ, время, особенности применения препарата: " + sheet['J' + str(u)].value
        self.textBrowser_disease_and_recom.setText(text_diseases)
        pass

    def search_type_bugs_set(self, name_pic):
        print("type")
        print("ok1")
        self.comboBox_all_bugs.clear()
        al = []
        print("ok2", al)
        with open("D:/BUGS_Programm/bin/info.json", "r") as read_file:
            info_pic = json.load(read_file)

        with open(name_pic, "r") as read_file:
            data_now_pic = json.load(read_file)
        mas = []
        self.comboBox_all_bugs.addItem(self.all)
        for i in range(len(data_now_pic["targets"])):
            for j in range(len(info_pic["classificator"])):
                mas.append(data_now_pic["targets"][i])
                if data_now_pic["targets"][i]["class_id"] == info_pic["classificator"][j]["id"] and data_now_pic["targets"][i]["class_id"] not in al:
                    al.append(data_now_pic["targets"][i]["class_id"])

                    self.comboBox_all_bugs.addItem(info_pic["classificator"][j]["name"])
                    print("added")
        #print(data_now_pic["targets"])

    def search_bugs_set(self):
        print(self.comboBox_all_bugs.currentText())
        self.listWidget_bugs.clear()
        l = []
        n = []
        k = 0
        if self.comboBox_all_bugs.currentText() == self.all:
            with open("D:/BUGS_Programm/bin/info.json", "r") as read_file:
                info_pic = json.load(read_file)
            for j in range(len(info_pic["classificator"])):
                l.append(info_pic["classificator"][j]["id"])
                n.append(info_pic["classificator"][j]["name"])
        else:
            l.append(self.comboBox_all_bugs.currentText())

        if self.comboBox_all_bugs.currentText() != "":
            for t in range(len(l)):
                k += 1
                if len(l) > 1:
                    self.type = l[t]
                    name = n[t]
                else:
                    print("ok33")
                    with open("D:/BUGS_Programm/bin/info.json", "r") as read_file:
                        info_pic = json.load(read_file)
                    print("ok34")
                    for i in range(len(info_pic["classificator"])):
                        print("text", self.comboBox_all_bugs.currentText())
                        if self.comboBox_all_bugs.currentText() == info_pic["classificator"][i]["name"]:
                            print("ok35")
                            self.type = info_pic["classificator"][i]["id"]
                            name = info_pic["classificator"][i]["name"]
                            print(self.type, name)
                            break

                with open(self.mydir_map + "/" + self.name_now_pic[:-4] + ".json", "r") as read_file:
                    data_now_pic = json.load(read_file)
                print("ok36")
                kk = 0
                for i in range(len(data_now_pic["targets"])):
                    if data_now_pic["targets"][i]["class_id"] == self.type:
                        kk += 1
                        self.listWidget_bugs.addItem(str(k) + "." + str(kk) + ":" + name)
                self.listWidget_bugs.itemClicked.connect(self.search_bugs)
                print("ok37")


    def search_bugs(self, item):
        print("search---")
        if self.comboBox_all_bugs.currentText() == "All":
            pass

        with open(self.mydir_map + "/" + self.name_now_pic[:-4] + ".json", "r") as read_file:
            data_now_pic = json.load(read_file)
        print(data_now_pic)
        print(format(item.text())[4:])

        with open("D:/BUGS_Programm/bin/info.json", "r") as read_file:
            info_pic = json.load(read_file)
        for j in range(len(info_pic["classificator"])):
            if info_pic["classificator"][j]["name"] == format(item.text())[4:]:
                self.type = info_pic["classificator"][j]["id"]
                print(self.type)
        s = int(format(item.text())[2])
        print(s, self.type)
        count = 0
        for i in range(len(data_now_pic["targets"])):
            if self.type == data_now_pic["targets"][i]["class_id"]:
                count += 1
                print(count)
                if count == s:
                    x = data_now_pic["targets"][i]["polygon"]

        print(x)
        cords = []
        num = ''
        for i in x:
            if i.isdigit():
                num += i
            else:
                if num != '':
                    cords.append(int(num))
                    num = ''
        print(cords)
        f = self.qpixmax.size().width() / ((cords[0] - cords[2]) * 10)
        self.scale = cords[4] - cords[0]
        self.viewer.fitInView()
        self.viewer.scale(f,f)
        self.viewer._zoom = 9
        self.viewer.centerOn(int((cords[0] + (cords[2] - cords[0]) / 2)), int((cords[1] + (cords[5] - cords[1]) / 2)))
        print("---")

    def map_coords(self, x, y, name):
        print("map_coords")
        wb = opxl.load_workbook(self.telemetry_path)
        sheet = wb['Лист1']

        coords = (float(sheet['B' + str(2)].value), float(sheet['C' + str(2)].value))
        self.m = folium.Map(
            title="lol",
            zoom_start=20,
            location=coords
        )
        path=[]
        path2=[]
        path3=[]
        print(sheet.max_row)
        for i in range(sheet.max_row):
            if sheet['B' + str(i + 2)].value == None:
                break
            print((float(sheet['B' + str(i + 2)].value), float(sheet['C' + str(i + 2)].value)))
            #path.append((int(sheet['B' + str(i + 2)].value), int(sheet['C' + str(i + 2)].value)))
            path.append((float(sheet['B' + str(i + 2)].value), float(sheet['C' + str(i + 2)].value)))
        print(path)
        folium.PolyLine(path,
                            color='red',
                            weight=3,
                            opacity=0.8).add_to(self.m)

        path2.append(path[36])
        path2.append(path[43])
        path2.append(path[48])
        path2.append(path[55])
        path2.append(path[60])
        path2.append(path[61])
        path2.append(path[66])
        path2.append(path[90])
        path2.append(path[36])
        folium.PolyLine(path2,
                        color='blue',
                        weight=2,
                        opacity=1).add_to(self.m)

        path3.append(path[79])
        path3.append(path[80])
        path3.append(path[81])
        path3.append(path[83])
        path3.append(path[84])
        path3.append(path[88])
        path3.append(path[90])
        path3.append(path[93])
        path3.append(path[79])
        folium.PolyLine(path3,
                        color='green',
                        weight=2,
                        opacity=1).add_to(self.m)
        print("ok")
      #for i in range(sheet.max_row):
        #    if sheet['B' + str(i + 2)].value == None:
        #        break
        #    folium.Marker([float(sheet['B' + str(i + 2)].value), float(sheet['C' + str(i + 2)].value)], popup=sheet['A' + str(i + 2)].value, tooltip=sheet['A' + str(i + 2)].value).add_to(self.m)
        folium.Marker([x, y], popup=name, tooltip=name, icon=folium.Icon(color='red')).add_to(self.m)
        data = io.BytesIO()
        self.m.save(data, close_file=False)

        webView = QWebEngineView()
        webView.setHtml(data.getvalue().decode())
        for i in reversed(range(self.verticalLayout_map.count())):
            self.verticalLayout_map.itemAt(i).widget().setParent(None)
        self.verticalLayout_map.addWidget(webView)
        print("end map_coords")


    def bugs_map(self):
        pass

    def changeEvent(self, event):
        if event.type() == QtCore.QEvent.LanguageChange:
            self.retranslateUi(self)
        super(MapWindow, self).changeEvent(event)
